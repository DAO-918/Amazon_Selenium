from numbers import Integral
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import pandas as pd
from Tools_Init import startInit
from Tools_Execl import pyxl_draw
from Tools_Win import get_files_by_name
from openpyxl.utils import get_column_letter
import os

## ! 步骤三从汇总拷贝数据到展示表

def excel_auto(file_name):
    config = startInit()
    '''
    链接	ASIN	国家	变体	品牌	店铺	备注    备注2	标题	AmazonChoice	图片1	图片2	图片3	图片4	图片5	图片6	图片7	图片8	视频	五点1	五点2	五点3	五点4	五点5	五点6 isDeal	现价	RRP	折扣	会员价	coupon	saving	promotion	基本信息	Keepa	卖家精灵	大类排名	小类排名	QA	评分	评价数	分数	特征
    这些是列名,请补充:column_name={
        'A':'链接'
    }
    '''
    column_name_dict = {
        'A': '链接',
        'B': 'ASIN',
        'C': '国家',
        'D': '变体',
        'E': '品牌',
        'F': '店铺',
        'G': '备注',
        'H': '备注2',
        'I': '标题',
        'J': 'AmazonChoice',
        'K': '图片1',
        'L': '图片2',
        'M': '图片3',
        'N': '图片4',
        'O': '图片5',
        'P': '图片6',
        'Q': '图片7',
        'R': '图片8',
        'S': '视频',
        'T': '五点1',
        'U': '五点2',
        'V': '五点3',
        'W': '五点4',
        'X': '五点5',
        'Y': '五点6',
        'Z': 'isDeal',
        'AA': '现价',
        'AB': 'RRP',
        'AC': '折扣',
        'AD': '会员价',
        'AE': 'coupon',
        'AF': 'saving',
        'AG': 'promotion',
        'AH': '基本信息',
        'AI': 'Keepa',
        'AJ': '卖家精灵',
        'AK': '大类排名',
        'AL': '小类排名',
        'AM': 'QA',
        'AN': '评分',
        'AO': '评价数',
        'AP': '分数',
        'AQ': '特征',
    }

    picture_file_path = config['picture_file_path']
    info_file_path = config['info_file_path']

    sheet_info_path = info_file_path + 'ASIN_Array_信息汇总.xlsx'
    sheet_info = pd.read_excel(sheet_info_path, sheet_name='Sheet1')

    sheet_ASIN_path = os.path.join('D:\\AutoRPA\\产品信息\\产品竞品', file_name)

    # 当data_only=True时，工作簿将只加载公式单元格的计算结果，而不加载公式本身
    # 如果需要修改或操作公式，使用默认的data_only=False参数来加载工作簿。
    wb = load_workbook(sheet_ASIN_path, data_only=False)
    sheet_ASIN = wb['Sheet1']
    # values_only=True的意义是在iter_rows方法中指定返回的行数据仅包含单元格的值，而不包括其他信息。这可以提高代码执行效率并简化对数据的访问。
    for row_index, row in enumerate(
        sheet_ASIN.iter_rows(min_row=2, values_only=True), start=2
    ):
        Link = row[0]
        ASIN = row[1]
        Country = row[2]
        if ASIN is None:
            break

        try:
            # 查找链接在ASIN_Array_信息汇总的行位置
            info_index = sheet_info.loc[sheet_info['链接'] == Link].index[0]
        except Exception:
            # 不能赋值为false，否则if flag >= 0: 会判断为true
            info_index = -1

        if info_index == -1:
            continue

        五点描述 = eval(sheet_info.at[info_index, '五点描述'])
        五点描述_len = len(五点描述)
        商品信息 = eval(sheet_info.at[info_index, '商品信息'])
        try:
            排名 = 商品信息['Best Sellers Rank'].split('\n', 1)
        except Exception:
            排名 = [None, None]

        # 图片
        #pic_floder = f'{picture_file_path}{ASIN}'
        pic_floder = os.path.join('D:\\AutoRPA\\产品图片', str(ASIN))
        pic_files = get_files_by_name(pic_floder, '450')
        if pic_files is not None:
            pyxl_draw(sheet_ASIN_path, wb, 'Sheet1', ASIN, pic_files, row_index, 11, min(len(pic_files), 8), 46, 8, False)  # type: ignore

        seller_canvas_path = f'D:\\AutoRPA\\卖家精灵\\{ASIN}\\seller_canvas.png'
        if os.path.exists(seller_canvas_path):
            pyxl_draw(sheet_ASIN_path, wb, 'Sheet1', ASIN, [seller_canvas_path], row_index, 36, 1, 46, 28, False)  # type: ignore

        seller_quickview_path = f'D:\\AutoRPA\\卖家精灵\\{ASIN}\\seller_quickview.png'
        if os.path.exists(seller_quickview_path):
            pyxl_draw(sheet_ASIN_path, wb, 'Sheet1', ASIN, [seller_quickview_path], row_index, 34, 1, 46, 15, False)  # type: ignore

        # enumerate(row, start=1) 创建一个迭代器，对于 row 列表中的每个元素，返回 (index, value) 对。start=1 参数设置起始索引值为 1，而不是默认的 0。
        # col_index 是当前元素在行中的索引，对应于工作表中的列索引。
        # cell 是当前元素的值，表示工作表中的单元格值。
        for col_index, cell in enumerate(row, start=1):
            # sheet_ASIN.cell(row=1, column=col_index).column_letter是使用openpyxl库中cell对象的column_letter属性来获取列名。这个属性直接返回列的字母表示，例如1对应'A'，2对应'B'，
            # column_letter = sheet_ASIN.cell(row=1, column=col_index).column_letter
            # get_column_letter(col_index)是使用openpyxl库中的get_column_letter函数来将列的序号转换为列名。该函数接受列的序号作为参数，并返回对应的列名。
            column_letter = get_column_letter(col_index)
            try:
                column_name = column_name_dict[column_letter]
                print(f'第{row_index}行，   正在写入第{col_index}列，   列名{column_name}')
            except Exception as e:
                print("超出列范围")
                break

            if column_name == 'ASIN':
                sheet_ASIN.cell(row=row_index, column=col_index).value = sheet_info.at[
                    info_index, 'ASIN'
                ]
            elif column_name == '国家':
                sheet_ASIN.cell(row=row_index, column=col_index).value = sheet_info.at[
                    info_index, '国家'
                ]
            elif column_name == '变体':
                sheet_ASIN.cell(row=row_index, column=col_index).value = sheet_info.at[
                    info_index, '变体'
                ]
            elif column_name == '品牌':
                sheet_ASIN.cell(row=row_index, column=col_index).value = str(
                    sheet_info.at[info_index, '品牌']
                )
            # 店铺
            elif column_name == '标题':
                sheet_ASIN.cell(row=row_index, column=col_index).value = sheet_info.at[
                    info_index, '标题'
                ]
            elif column_name == 'AmazonChoice':
                sheet_ASIN.cell(row=row_index, column=col_index).value = sheet_info.at[
                    info_index, 'AmazonChoice'
                ]
            elif column_name == '视频数量':
                sheet_ASIN.cell(row=row_index, column=col_index).value = sheet_info.at[
                    info_index, '视频数量'
                ]
            elif column_name == '五点1' and 五点描述_len >= 1:
                sheet_ASIN.cell(row=row_index, column=col_index).value = 五点描述[0]
            elif column_name == '五点2' and 五点描述_len >= 2:
                sheet_ASIN.cell(row=row_index, column=col_index).value = 五点描述[1]
            elif column_name == '五点3' and 五点描述_len >= 3:
                sheet_ASIN.cell(row=row_index, column=col_index).value = 五点描述[2]
            elif column_name == '五点4' and 五点描述_len >= 4:
                sheet_ASIN.cell(row=row_index, column=col_index).value = 五点描述[3]
            elif column_name == '五点5' and 五点描述_len >= 5:
                sheet_ASIN.cell(row=row_index, column=col_index).value = 五点描述[4]
            elif column_name == '五点6' and 五点描述_len >= 6:
                sheet_ASIN.cell(row=row_index, column=col_index).value = 五点描述[5]
            elif column_name == 'isDeal':
                sheet_ASIN.cell(row=row_index, column=col_index).value = bool(
                    sheet_info.at[info_index, 'isDeal']
                )
            elif column_name == '现价':
                sheet_ASIN.cell(row=row_index, column=col_index).value = sheet_info.at[
                    info_index, '现价'
                ]
            elif column_name == 'RRP':
                sheet_ASIN.cell(row=row_index, column=col_index).value = sheet_info.at[
                    info_index, 'RRP'
                ]
            elif column_name == '折扣':
                sheet_ASIN.cell(row=row_index, column=col_index).value = sheet_info.at[
                    info_index, '折扣'
                ]
            elif column_name == '会员价':
                sheet_ASIN.cell(row=row_index, column=col_index).value = sheet_info.at[
                    info_index, '会员价'
                ]
            elif column_name == 'coupon':
                sheet_ASIN.cell(row=row_index, column=col_index).value = sheet_info.at[
                    info_index, 'coupon'
                ]
            elif column_name == 'saving':
                sheet_ASIN.cell(row=row_index, column=col_index).value = sheet_info.at[
                    info_index, 'saving'
                ]
            elif column_name == 'promotion':
                sheet_ASIN.cell(row=row_index, column=col_index).value = sheet_info.at[
                    info_index, 'promotion'
                ]
            elif column_name == '大类排名':
                sheet_ASIN.cell(row=row_index, column=col_index).value = 排名[0]
            elif column_name == '小类排名':
                sheet_ASIN.cell(row=row_index, column=col_index).value = 排名[1]
            elif column_name == '评分':
                sheet_ASIN.cell(row=row_index, column=col_index).value = sheet_info.at[
                    info_index, '评分'
                ]
            elif column_name == '评价数':
                sheet_ASIN.cell(row=row_index, column=col_index).value = sheet_info.at[
                    info_index, '评价数'
                ]

    # sheet_ASIN_path = f'{config["info_file_path"]} 产品竞品\\{ASIN_Info-TOP22003_2.xlsx}'
    
    flag = True
    # 设置自动换行的列
    if flag == True:
        columns_to_wrap = [
            'A',
            'E',
            'G',
            'H',
            'I',
            'J',
            'T',
            'U',
            'V',
            'W',
            'X',
            'Y',
            'AK',
            'AL',
        ]
        for column in columns_to_wrap:
            for cell in sheet_ASIN[column]:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # 设置字体大小的列
        columns_size_9 = ['S', 'T', 'U', 'V', 'W', 'X', 'Y']
        font_size = 9
        for column in columns_size_9:
            for cell in sheet_ASIN[column]:
                cell.font = Font(size=font_size)

        # 设置None替换为空
        for row in sheet_ASIN.rows:
            for cell in row:
                if cell.value == 'None':
                    cell.value = ''

        # 设置列宽
        sheet_ASIN.column_dimensions['B'].width = 10
        sheet_ASIN.column_dimensions['C'].width = 4
        sheet_ASIN.column_dimensions['D'].width = 6
        sheet_ASIN.column_dimensions['I'].width = 10

    wb.save(sheet_ASIN_path)


if __name__ == '__main__':
    #excel_auto('ASIN_Info-测试.xlsx')
    excel_auto('ASIN_Info-总asin.xlsx')
