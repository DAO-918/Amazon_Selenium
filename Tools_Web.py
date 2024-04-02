from calendar import c
import os
import string
import time
import re
import json
from tkinter import N
import yaml
from datetime import datetime

import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions
from bs4 import BeautifulSoup

import urllib.parse
# from urllib.parse import urlparse
import logging

from collections import defaultdict

from docx import Document

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
import requests
import threading

from io import BytesIO
from PIL import Image

def picdownload(us_cookies, uk_cookies, url, asin, country, img_list: list, image_dir, filename):
    '''options = webdriver.ChromeOptions()
    options.binary_location = r'D:\Code\chrome-win\chrome.exe'
    options.debugger_address = '127.0.0.1:9222'
    options.browser_version = '114.0.5734.0'
    #禁用Chrome浏览器的自动控制提示。当使用selenium开启Chrome浏览器时，浏览器的信息栏上方会有一条提示：“Chrome正在受到自动测试软件的控制”
    #options.add_experimental_option("excludeSwitches", ["enable-automation"])
    #禁用Selenium提供的自动化扩展。Selenium默认会载入一些用于自动化控制的浏览器扩展
    #options.add_experimental_option('useAutomationExtension', False)
    service = Service(executable_path=r'D:\Code\chromedriver_win32\114\chromedriver.exe')
    driver = webdriver.Chrome(service=service, options=options)'''
    '''driver.get("https://www.amazon.com")
    us_cookies = driver.get_cookies()
    driver.get("https://www.amazon.co.uk")
    uk_cookies = driver.get_cookies()'''

    lock = threading.Lock()  # 创建一个锁对象
    xlsx_file = r'D:/Code/# LISTING/产品图片/ASIN_图片链接_450.xlsx'

    def download(pic_url, file_path, cookie, row, column):
        RETRY_TIMES = 5
        RETRY_INTERVAL = 2
        headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'zh-CN,zh-Hans;q=0.9',
            'Connection': 'keep-alive',
            'Cookie':'; '.join([item["name"]+"="+item["value"] for item in cookie]),
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
            }
        for i in range(RETRY_TIMES):
            try:
                # make request with headers and cookies
                res = response = requests.get(pic_url, headers=headers, timeout=5)
                print(res.status_code)
                # https://blog.csdn.net/wantpython_king/article/details/132185135
                # 在使用requests请求网页，并将图片保存到本地时，然后通过openyxl插入到excel，在最后的wb.save()时报错keyError: ‘.webp’，通过筛选找到报错的图片，发现本地有这个图片并且能正常打开，就是不能正常插入到excel中。
                with open(file_path, 'wb') as f:
                    # write content to file
                    f.write(response.content)
                if res.status_code == 200:
                    img = Image.open(BytesIO(res.content))
                    img.save(file_path)
                # mark as downloaded in excel
                with lock:  # 确保线程安全
                    ws.cell(row=row, column=column, value='TRUE')
                break
            except requests.exceptions.HTTPError as errh:
                print ("Http Error:",errh)
            except requests.exceptions.ConnectionError as errc:
                print ("Error Connecting:",errc)
            except requests.exceptions.Timeout as errt:
                print ("Timeout Error:",errt)
            except requests.exceptions.RequestException as err:
                print ("Req Error:",err)
            except Exception as err:
                print ("?Exception:",err)
            finally:
                print(f"Download failed, retrying {i+1}/{RETRY_TIMES}")
                time.sleep(1)
        else:
            print(f"All attempts to download have failed: {pic_url}")
            # mark as download failed in excel
            with lock:  # 确保线程安全
                ws.cell(row=row, column=column, value='FALSE')
        '''RETRY_TIMES = 5
        RETRY_INTERVAL = 1
        for i in range(RETRY_TIMES):
            try:
                headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36'}
                response = requests.get(pic_url,headers=headers, timeout=5)
                with open(file_path, 'wb') as f:
                    f.write(response.content)
                with lock:  # 确保线程安全
                    ws.cell(row=row, column=column, value='TRUE')
                break
            except Exception:
                print(f"Download failed, retrying {i+1}/{RETRY_TIMES}")
                time.sleep(RETRY_INTERVAL)  
        else:
            print(f"All attempts to download have failed: {pic_url}")
            with lock:  # 确保线程安全
                ws.cell(row=row, column=column, value='FALSE')'''

    # 加载xlsx文件
    wb = load_workbook(filename=xlsx_file)
    ws = wb['Sheet1']

    
    picture_file_path = r'D:\AutoRPA\产品图片'
    # 找出链接列
    columns = [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    link_col = columns.index('链接')

    # 搜索对应链接
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True),start=2):
        if row[link_col] == url:
            matched_row = row
            matched_row_index = row_index
            break
    else:
        matched_row = None

    if matched_row is None:  # 创建新行
        new_row = ['' for _ in range(len(columns))]
        new_row[link_col] = url
        new_row[columns.index('ASIN')] = asin
        new_row[columns.index('国家')] = country
        for i, img_link in enumerate(img_list, 1):
            new_row[columns.index(f'image_450_{i}')] = img_link
            new_row[columns.index(f'image_450_{i}_download')] = 'FALSE'
        ws.append(new_row)
    else:  # 更新已有行
        for i, img_link in enumerate(img_list, 1):
            if matched_row[columns.index(f'image_450_{i}')] != img_link:
                ws.cell(row=matched_row_index, column=columns.index(f'image_450_{i}')+1, value=img_link)
                ws.cell(row=matched_row_index, column=columns.index(f'image_450_{i}_download')+1, value='FALSE')

    # 保存xlsx文件
    wb.save(xlsx_file)
    
    columns = [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    threads = []  # 用来保存启动的线程

    # loop through each row in the excel file
    for row_index, row in enumerate(ws.iter_rows(min_row=2),start=2):
        # get country and img_link
        url = row[0].value
        asin = row[1].value
        country = row[2].value
        image_dir = os.path.join(picture_file_path, str(asin))
        if not os.path.exists(image_dir):
            os.makedirs(image_dir)
        #cookie = cookies_dict[str(country).lower()]
        if str(country).lower() == 'us':
            cookie = us_cookies
        elif str(country).lower() == 'uk':
            cookie = uk_cookies
        for i in range(1,10):
            image_url_column = columns.index(f'image_450_{i}') + 1
            image_url = ws.cell(row=row_index, column=image_url_column).value
            image_download_column = columns.index(f'image_450_{i}_download')+1
            image_download = ws.cell(row=row_index, column=image_download_column).value
            #if image_download == False:
            #    file_path = os.path.join(image_dir, f'image_450_{i}.jpg')
            #    t = threading.Thread(target=download, args=(image_url, file_path, cookie, row_index, image_download_column))
            #    threads.append(t)
            #    t.start()

    '''# 下载图片并更新状态
    threads = []  # 用来保存启动的线程
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True),start=2):
        if row[link_col] == url:
            for i, img_link in enumerate(img_list, 1):
                if row[columns.index(f'image_450_{i}_download')] == 'FALSE':
                    file_path = os.path.join(image_dir, f'image_450_{i}.jpg')
                    t = threading.Thread(target=download, args=(img_link, file_path, cookie, row_index, columns.index(f'image_450_{i}_download')+1))
                    threads.append(t)
                    t.start()'''
                    #urllib.request.urlretrieve(img_link, file_path)
                    #ws.cell(row=row_index, column=columns.index(f'image_450_{i}_download')+1, value='TRUE')
        
    #for thread in threads:
    #    thread.join()  # 等待所有线程结束

    # 保存xlsx文件
    wb.save(xlsx_file)

# 测试代码
if __name__ == '__main__':
    # 准备工作
    img_list = ['https://m.media-amazon.com/images/S/aplus-media-library-service-media/bc65a8d0-e7e6-4ae0-b537-a8a35bb65ead.__CR0,0,970,600_PT0_SX970_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/0582c7d9-3b4e-47f0-aaf4-10efa7d6f629.__CR0,0,300,400_PT0_SX300_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/ff0c023e-f208-41b6-b4fb-d507f37408d3.__CR0,0,300,300_PT0_SX300_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/5220da0f-ba34-4ebb-bb5a-3347844eb832.__CR0,0,350,175_PT0_SX350_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/3f2897b5-d435-4a94-a1ee-5d9abafeaf54.__CR0,0,970,600_PT0_SX970_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/d08b3044-d5a0-447c-a614-a6ecab4f9f10.__CR0,0,970,300_PT0_SX970_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/f29d9999-1903-4d66-a083-2b11fa542701.__CR0,0,300,300_PT0_SX300_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/7701fa4e-e967-425f-8780-24a789ee0406.__CR0,0,970,600_PT0_SX970_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/d6b01f61-da01-49b7-918a-e7f80ab4fde4.__CR0,0,970,600_PT0_SX970_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/b1e1ad4c-5708-422b-ad8f-8cedf3a0bb74.__CR0,0,300,300_PT0_SX300_V1___.jpg']
    url = 'https://www.amazon.com/dp/B0C6993ST3'
    asin = 'B0C6993ST3'
    picture_file_path = r'D:\AutoRPA\产品图片'
    image_dir = f'{picture_file_path}/{asin}/'
    country = 'US'
    #picdownload(url, asin, country, img_list, image_dir, '450')
