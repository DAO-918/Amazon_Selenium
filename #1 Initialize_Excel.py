import re
from Tools_Init import startInit
from Tools_Execl import link_AutoComple
import pandas as pd
from openpyxl import load_workbook

import urllib.parse

## ! 步骤一格式化链接
config = startInit()

def format_url(url:str):
    if 'aax-us-iad.' in url:
        url = url.replace('aax-us-iad', 'www')
    if 'sspa' in url and 'dp%2F' in url:
        asin_pattern = re.compile(r'dp%2F([A-Za-z0-9]{10})')
    elif 'www.amazon.com' in url or 'www.amazon.co.uk' in url:
        asin_pattern = re.compile(r'/dp/([A-Z0-9]{10})')
    else:
        return None, None, None
    match = asin_pattern.search(url)
    asin = match[1] if match else None
    parsed_url = urllib.parse.urlparse(url)
    domain = parsed_url.netloc
    domain_parts = domain.split('.')
    if len(domain_parts) >= 4:
        domain_suffix = f'{domain_parts[-2]}.{domain_parts[-1]}'
    else:
        domain_suffix = domain_parts[-1]
    domain_suffix_country_dict = {'com': 'us', 'co.uk': 'uk'}
    country = domain_suffix_country_dict.get(domain_suffix, domain_suffix)
    new_url = f'https://{domain}/dp/{asin}'
    return new_url, asin, country

def Initialize_Excel(sheet_ASIN_path):
    sheet_ASIN = pd.read_excel(sheet_ASIN_path, sheet_name='Sheet1')
    wb = load_workbook(sheet_ASIN_path, data_only=False)
    sheet_ASIN = wb['Sheet1']

    for row_index, row in enumerate( # type: ignore
        sheet_ASIN.iter_rows(min_row=2, values_only=True), start=2
    ):    
        link = str(row[0])
        asin = str(row[1])
        country = str(row[2])
        if link is None or link=='None':
            break
        '''# 如果链接不为空
        #if not pd.isna(link) and pd.isna(asin) and pd.isna(country):
        url_array = link.split('/')
        # 编译正则表达式提取ASIN
        asin_pattern = re.compile(r'/dp/([A-Z0-9]{10})')

        # 从url1提取ASIN
        match = asin_pattern.search(link)
        ASIN = match[1] if match else None  

        # 从url提取国家代码
        Country = url_array[2].split('.')[-1]
        if Country == 'com':
            Country = 'us'
        elif Country == 'co.uk':
            Country = 'uk'''
        '''#elif pd.isna(link) and not pd.isna(asin) and not pd.isna(country):
        if Country == 'us':
            Link = f'https://www.amazon.com/dp/{ASIN}'
        elif Country == 'uk':
            Link = f'https://www.amazon.co.uk/dp/{ASIN}'
        else:
            Link = f'https://www.amazon.{Country}/dp/{ASIN}'
            # .at[] 是针对单个元素进行赋值的方法，只能更新一个单元格的值。
            # .loc[] 是针对切片进行赋值的方法，可以同时更新多个单元格的值。'''
        
        Link,ASIN,Country = format_url(link)
        if Link is None:
            sheet_ASIN.cell(row=row_index, column=4).value = 'link error' # type: ignore
            continue
        sheet_ASIN.cell(row=row_index, column=1).value = Link # type: ignore
        sheet_ASIN.cell(row=row_index, column=2).value = ASIN # type: ignore
        sheet_ASIN.cell(row=row_index, column=3).value = Country # type: ignore

    wb.save(sheet_ASIN_path)

# 测试代码
if __name__ == '__main__':
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\ASIN_Array_抓取队列.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-遥控双轮特技车.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-益智华容道.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-电子魔方.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-遥控越野.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-飞碟发射器.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-射击靶空气枪.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-暂存.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-传统蒸汽火车.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-物理科学实验套装.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-遥控翻斗车.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-恐龙卡车.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-大型维修套装.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-HL808迭代.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-玩偶小屋.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-乐高花束.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-高级电子教学.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-Wal.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-汇乐.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-暂存3.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-带射击特技车.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-电子绘本.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-读卡器.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-APP机器人.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-圣诞树.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-ZM21001.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-粘球枪.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-螃蟹.xlsx')
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-乐高花束2.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-总asin.xlsx')