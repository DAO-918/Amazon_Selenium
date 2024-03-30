import re
from Tools_Init import startInit
from Tools_Execl import link_AutoComple
import pandas as pd
from openpyxl import load_workbook

config = startInit()

def Initialize_Excel(sheet_ASIN_path):
    sheet_ASIN = pd.read_excel(sheet_ASIN_path, sheet_name='Sheet1')
    wb = load_workbook(sheet_ASIN_path, data_only=False)
    sheet_ASIN = wb['Sheet1']

    for row_index, row in enumerate(
        sheet_ASIN.iter_rows(min_row=2, values_only=True), start=2
    ):    
        link = str(row[0])
        asin = str(row[1])
        country = str(row[2])
        if link is None or link=='None':
            break
        # 如果链接不为空
        #if not pd.isna(link) and pd.isna(asin) and pd.isna(country):
        url_array = link.split('/')
        # 编译正则表达式提取ASIN
        asin_pattern = re.compile(r'/dp/([A-Z0-9]{10})')

        # 从url1提取ASIN
        match = asin_pattern.search(link)
        ASIN = match[1] if match else None  

        # 从url提取国家代码
        Country = url_array[2].split('.')[-1]
        if Country == 'com':
            Country = 'us'
        elif Country == 'co.uk':
            Country = 'uk'

        #elif pd.isna(link) and not pd.isna(asin) and not pd.isna(country):
        if Country == 'us':
            Link = f'https://www.amazon.com/dp/{ASIN}'
        elif Country == 'uk':
            Link = f'https://www.amazon.co.uk/dp/{ASIN}'
        else:
            Link = f'https://www.amazon.{Country}/dp/{ASIN}'
            # .at[] 是针对单个元素进行赋值的方法，只能更新一个单元格的值。
            # .loc[] 是针对切片进行赋值的方法，可以同时更新多个单元格的值。

        sheet_ASIN.cell(row=row_index, column=1).value = Link # type: ignore
        sheet_ASIN.cell(row=row_index, column=2).value = ASIN # type: ignore
        sheet_ASIN.cell(row=row_index, column=3).value = Country # type: ignore

    wb.save(sheet_ASIN_path)

# 测试代码
if __name__ == '__main__':
    #Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-暂存.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-遥控双轮特技车.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-遥控两栖.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-遥控翻斗车.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-遥控小车.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-遥控车.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-存钱罐.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-遥控特技.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-存钱罐.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-多米诺火车.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-恐龙手部装扮.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-忙碌电子板.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-忙碌机关板.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-忙碌毛毡书.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-汽车闯关大冒险.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-高速轨道.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-新抓取轨道.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-重力百变轨道.xlsx')
    Initialize_Excel(r'D:\AutoRPA\产品信息\ASIN_Array_抓取队列.xlsx')
