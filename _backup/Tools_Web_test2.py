from openpyxl import load_workbook
import os
import urllib

# 准备工作
img_list = ['https://m.media-amazon.com/images/S/aplus-media-library-service-media/bc65a8d0-e7e6-4ae0-b537-a8a35bb65ead.__CR0,0,970,600_PT0_SX970_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/0582c7d9-3b4e-47f0-aaf4-10efa7d6f629.__CR0,0,300,400_PT0_SX300_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/ff0c023e-f208-41b6-b4fb-d507f37408d3.__CR0,0,300,300_PT0_SX300_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/5220da0f-ba34-4ebb-bb5a-3347844eb832.__CR0,0,350,175_PT0_SX350_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/3f2897b5-d435-4a94-a1ee-5d9abafeaf54.__CR0,0,970,600_PT0_SX970_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/d08b3044-d5a0-447c-a614-a6ecab4f9f10.__CR0,0,970,300_PT0_SX970_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/f29d9999-1903-4d66-a083-2b11fa542701.__CR0,0,300,300_PT0_SX300_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/7701fa4e-e967-425f-8780-24a789ee0406.__CR0,0,970,600_PT0_SX970_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/d6b01f61-da01-49b7-918a-e7f80ab4fde4.__CR0,0,970,600_PT0_SX970_V1___.jpg', 'https://m.media-amazon.com/images/S/aplus-media-library-service-media/b1e1ad4c-5708-422b-ad8f-8cedf3a0bb74.__CR0,0,300,300_PT0_SX300_V1___.jpg']
url = 'https://www.amazon.com/dp/B0C6993ST3'
asin = 'B0C6993ST3'
picture_file_path = r'D:\AutoRPA\产品图片'
image_dir = f'{picture_file_path}/{asin}/'
xlsx_file = r'D:/Code/# LISTING/产品图片/ASIN_图片链接_450.xlsx'
country = 'US'

# 加载xlsx文件
wb = load_workbook(xlsx_file)
ws = wb.active

# 找出链接列
columns = [cell.value for cell in ws[1]]
link_col = columns.index('链接')

# 搜索对应链接
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[link_col] == url:
        matched_row = row
        break
else:
    matched_row = None

if matched_row is None:  # 创建新行
    new_row = [None] * len(columns)
    new_row[link_col] = url
    new_row[columns.index('ASIN')] = asin
    new_row[columns.index('国家')] = country
    for i, img_link in enumerate(img_list, 1):
        new_row[columns.index(f'image_450_{i}')] = img_link
        new_row[columns.index(f'image_450_{i}_download')] = 'FALSE'
    ws.append(new_row)
else:  # 更新已有行
    for i, img_link in enumerate(img_list, 1):
        if matched_row[columns.index(f'image_450_{i}')] != img_link:
            ws.cell(row=matched_row[0].row, column=columns.index(f'image_450_{i}')+1, value=img_link)
            ws.cell(row=matched_row[0].row, column=columns.index(f'image_450_{i}_download')+1, value='FALSE')

# 保存xlsx文件
wb.save(xlsx_file)

# 重新加载xlsx文件
wb = load_workbook(xlsx_file)
ws = wb.active

# 下载图片并更新状态
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[link_col] == url:
        for i, img_link in enumerate(img_list, 1):
            if row[columns.index(f'image_450_{i}_download')] == 'FALSE':
                file_path = os.path.join(image_dir, f'image_450_{i}.jpg')
                urllib.request.urlretrieve(img_link, file_path)
                ws.cell(row=row[0].row, column=columns.index(f'image_450_{i}_download')+1, value='TRUE')
        break

# 保存xlsx文件
wb.save(xlsx_file)
