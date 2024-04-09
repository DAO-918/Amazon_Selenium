from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions

from io import BytesIO

options = webdriver.ChromeOptions()
options.binary_location = r'D:\Code\chrome-win\chrome.exe'
options.debugger_address = '127.0.0.1:9222'
options.browser_version = '114.0.5734.0'
#禁用Chrome浏览器的自动控制提示。当使用selenium开启Chrome浏览器时，浏览器的信息栏上方会有一条提示：“Chrome正在受到自动测试软件的控制”
#options.add_experimental_option("excludeSwitches", ["enable-automation"])
#禁用Selenium提供的自动化扩展。Selenium默认会载入一些用于自动化控制的浏览器扩展
#options.add_experimental_option('useAutomationExtension', False)
service = Service(executable_path=r'D:\Code\chromedriver_win32\114\chromedriver.exe')
driver = webdriver.Chrome(service=service, options=options)

'''# to store cookies for each country's amazon website
cookies_dict = {} 
flag = False
def get_cookies(domain):
    # visit the domain
    driver.get(domain)
    # get the cookies
    cookies = driver.get_cookies()
    # store in the dictionary
    if domain == "https://www.amazon.com":
        cookies_dict['us'] = cookies 
    if domain == "https://www.amazon.co.uk":
        cookies_dict['uk'] = cookies 

if flag == True:
    driver.get("https://www.amazon.com")
    wait = WebDriverWait(driver, 60)
    wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'body')))
get_cookies("https://www.amazon.com")

if flag == True:
    driver.get("https://www.amazon.co.uk")
    wait = WebDriverWait(driver, 60)
    wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'body')))
get_cookies("https://www.amazon.co.uk")'''
flag = False
if flag == True:
    driver.get("https://www.amazon.com")
    wait = WebDriverWait(driver, 60)
    wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'body')))
if flag == True:
    driver.get("https://www.amazon.co.uk")
    wait = WebDriverWait(driver, 60)
    wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'body')))
driver.get("https://www.amazon.com")
us_cookies = driver.get_cookies()
driver.get("https://www.amazon.co.uk")
uk_cookies = driver.get_cookies()

# 以chrome为例,在地址栏输入about:version

driver.quit()

from openpyxl import load_workbook
import os
import time
import requests
import threading
from PIL import Image

# 加载xlsx文件
xlsx_file = r'D:\Code\# LISTING\产品图片\ASIN_图片链接_450.xlsx'
wb = load_workbook(filename=xlsx_file)  # replace 'excel_file_path' with your excel file path
ws = wb['Sheet1']

lock = threading.Lock()  # 创建一个锁对象

# image download function
def download(pic_url, file_path, cookie, row, column):
    RETRY_TIMES = 5
    RETRY_INTERVAL = 2
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh-Hans;q=0.9',
        'Connection': 'keep-alive',
        'Cookie':'; '.join([item["name"]+"="+item["value"] for item in cookie]),
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
        }
    for i in range(RETRY_TIMES):
        try:
            # make request with headers and cookies
            res = response = requests.get(pic_url, headers=headers, timeout=20)
            #print(res.status_code)
            # https://blog.csdn.net/wantpython_king/article/details/132185135
            # 在使用requests请求网页，并将图片保存到本地时，然后通过openyxl插入到excel，在最后的wb.save()时报错keyError: ‘.webp’，通过筛选找到报错的图片，发现本地有这个图片并且能正常打开，就是不能正常插入到excel中。
            with open(file_path, 'wb') as f:
                # write content to file
                f.write(response.content)
            if res.status_code == 200:
                img = Image.open(BytesIO(res.content))
                img.save(file_path)
            # mark as downloaded in excel
            with lock:  # 确保线程安全
                ws.cell(row=row, column=column, value='TRUE')
            break
        except requests.exceptions.HTTPError as errh:
            print ("Http Error:",errh)
        except requests.exceptions.ConnectionError as errc:
            print ("Error Connecting:",errc)
        except requests.exceptions.Timeout as errt:
            print ("Timeout Error:",errt)
        except requests.exceptions.RequestException as err:
            print ("Req Error:",err)
        except Exception as err:
            print ("?Exception:",err)
        finally:
            print(f"Download failed, retrying {i+1}/{RETRY_TIMES}")
            time.sleep(1)
    else:
        print(f"All attempts to download have failed: {pic_url}")
        # mark as download failed in excel
        with lock:  # 确保线程安全
            ws.cell(row=row, column=column, value='FALSE')

picture_file_path = r'D:\AutoRPA\产品图片'
columns = [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]
threads = []  # 用来保存启动的线程

# loop through each row in the excel file
for row_index, row in enumerate(ws.iter_rows(min_row=2),start=2):
    # get country and img_link
    url = row[0].value
    asin = row[1].value
    country = row[2].value
    image_dir = os.path.join(picture_file_path, str(asin))
    if not os.path.exists(image_dir):
        os.makedirs(image_dir)
    #cookie = cookies_dict[str(country).lower()]
    if str(country).lower() == 'us':
        cookie = us_cookies
    elif str(country).lower() == 'uk':
        cookie = uk_cookies
    for i in range(1,10):
        image_url_column = columns.index(f'image_450_{i}') + 1
        image_url = ws.cell(row=row_index, column=image_url_column).value
        image_download_column = columns.index(f'image_450_{i}_download')+1
        image_download = ws.cell(row=row_index, column=image_download_column).value
        if image_download == 'FALSE':
            file_path = os.path.join(image_dir, f'image_450_{i}.jpg')
            t = threading.Thread(target=download, args=(image_url, file_path, cookie, row_index, image_download_column))
            threads.append(t)
            t.start()

for thread in threads:
    thread.join()  # 等待所有线程结束

# 保存xlsx文件
wb.save(xlsx_file)